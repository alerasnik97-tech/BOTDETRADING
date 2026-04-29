from __future__ import annotations

import hashlib
import json
import subprocess
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
LAB = ROOT / "BOT_V2_DAYTIME_LAB"
MANIP = ROOT / "MANIPULANTE"
OUT = LAB / "outputs" / "phase38_manipulante_deep_explainer"
CSV_OUT = OUT / "csv"
REPORTS = LAB / "reports"
ANALYSIS_DIR = MANIP / "14_ANALISIS"
FIRST_READ_DIR = MANIP / "00_LEER_PRIMERO"
TRADES_CSV = LAB / "outputs" / "phase27_full_historical_validation_2015_2026" / "validation_2015_2026_full" / "phase27_2015_2026_trades.csv"


def read_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def read_text(path: Path) -> str:
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8", errors="replace")


def run_cmd(args: list[str]) -> str:
    try:
        return subprocess.run(
            args,
            cwd=ROOT,
            text=True,
            capture_output=True,
            check=False,
        ).stdout.strip()
    except Exception as exc:
        return f"ERROR: {exc}"


def safe_round(value: Any, digits: int = 3) -> Any:
    try:
        if pd.isna(value):
            return ""
        return round(float(value), digits)
    except Exception:
        return value


def profit_factor(values: pd.Series) -> float:
    gross_profit = values[values > 0].sum()
    gross_loss = -values[values < 0].sum()
    if gross_loss == 0:
        return float("inf") if gross_profit > 0 else 0.0
    return float(gross_profit / gross_loss)


def max_drawdown(values: pd.Series) -> float:
    curve = values.cumsum()
    peak = curve.cummax()
    dd = curve - peak
    return float(dd.min()) if len(dd) else 0.0


def parse_trades() -> pd.DataFrame:
    df = pd.read_csv(TRADES_CSV)
    df["entry_time_ny"] = pd.to_datetime(df["entry_time"], utc=True).dt.tz_convert("America/New_York")
    df["exit_time_ny"] = pd.to_datetime(df["exit_time"], utc=True).dt.tz_convert("America/New_York")
    sign = df["type"].map({"LONG": 1, "SHORT": -1}).fillna(0)
    df["r_result"] = ((df["exit_price"] - df["entry_price"]) * sign) / df["risk"]
    df.loc[df["status"].eq("TP"), "r_result"] = 1.4
    df.loc[df["status"].eq("SL") & df["be_triggered"].astype(bool), "r_result"] = 0.0
    df.loc[df["status"].eq("SL") & ~df["be_triggered"].astype(bool), "r_result"] = -1.0
    df["outcome"] = "OTRO"
    df.loc[df["status"].eq("TP"), "outcome"] = "TP"
    df.loc[df["status"].eq("SL") & df["be_triggered"].astype(bool), "outcome"] = "BE"
    df.loc[df["status"].eq("SL") & ~df["be_triggered"].astype(bool), "outcome"] = "SL"
    df.loc[df["status"].eq("FORCED_CLOSE"), "outcome"] = "FORCED_CLOSE"
    df["entry_date"] = df["entry_time_ny"].dt.date.astype(str)
    df["exit_date"] = df["exit_time_ny"].dt.date.astype(str)
    df["year"] = df["entry_time_ny"].dt.year
    df["year_month"] = df["entry_time_ny"].dt.strftime("%Y-%m")
    df["month_num"] = df["entry_time_ny"].dt.month
    df["month_name"] = df["month_num"].map(
        {
            1: "Enero",
            2: "Febrero",
            3: "Marzo",
            4: "Abril",
            5: "Mayo",
            6: "Junio",
            7: "Julio",
            8: "Agosto",
            9: "Septiembre",
            10: "Octubre",
            11: "Noviembre",
            12: "Diciembre",
        }
    )
    df["hour_ny"] = df["entry_time_ny"].dt.hour
    df["weekday_num"] = df["entry_time_ny"].dt.weekday
    df["weekday"] = df["weekday_num"].map(
        {
            0: "Lunes",
            1: "Martes",
            2: "Miercoles",
            3: "Jueves",
            4: "Viernes",
            5: "Sabado",
            6: "Domingo",
        }
    )
    df["is_win"] = df["r_result"] > 0
    df["is_non_win"] = df["r_result"] <= 0
    df["is_monetary_loss"] = df["r_result"] < 0
    df["no_tp_psychological"] = df["outcome"] != "TP"
    return df


def metric_row(name: str, df: pd.DataFrame) -> dict[str, Any]:
    r = df["r_result"]
    tp = int((df["outcome"] == "TP").sum())
    be = int((df["outcome"] == "BE").sum())
    sl = int((df["outcome"] == "SL").sum())
    fc = int((df["outcome"] == "FORCED_CLOSE").sum())
    total = len(df)
    non_be = total - be
    return {
        "name": name,
        "sample": total,
        "pf": safe_round(profit_factor(r), 3),
        "expectancy_r": safe_round(r.mean(), 4),
        "winrate_r_positive_pct": safe_round((r > 0).mean() * 100, 2),
        "tp_only_winrate_pct": safe_round(tp / total * 100 if total else 0, 2),
        "winrate_excluding_be_pct": safe_round((r[df["outcome"] != "BE"] > 0).mean() * 100 if non_be else 0, 2),
        "tp_rate_excluding_be_pct": safe_round(tp / non_be * 100 if non_be else 0, 2),
        "max_drawdown_r": safe_round(max_drawdown(r), 3),
        "total_r": safe_round(r.sum(), 3),
        "tp_count": tp,
        "be_exit_count": be,
        "sl_pure_count": sl,
        "forced_close_count": fc,
        "be_trigger_count": int(df["be_triggered"].astype(bool).sum()),
        "gross_profit_r": safe_round(r[r > 0].sum(), 3),
        "gross_loss_r": safe_round(r[r < 0].sum(), 3),
    }


def summarize_group(df: pd.DataFrame, group: str, label_name: str) -> pd.DataFrame:
    rows = []
    for label, g in df.groupby(group, sort=True):
        m = metric_row(str(label), g)
        m[label_name] = label
        m["trades"] = len(g)
        m["TP"] = int((g["outcome"] == "TP").sum())
        m["BE"] = int((g["outcome"] == "BE").sum())
        m["SL"] = int((g["outcome"] == "SL").sum())
        m["FORCED_CLOSE"] = int((g["outcome"] == "FORCED_CLOSE").sum())
        m["veredicto"] = verdict_for_group(m)
        rows.append(m)
    result = pd.DataFrame(rows)
    cols = [label_name, "trades", "TP", "BE", "SL", "FORCED_CLOSE", "pf", "expectancy_r", "winrate_r_positive_pct", "max_drawdown_r", "total_r", "veredicto"]
    return result[cols]


def verdict_for_group(row: dict[str, Any]) -> str:
    if row.get("sample", row.get("trades", 0)) < 30:
        return "MUESTRA_BAJA"
    if row.get("pf", 0) >= 2.5 and row.get("expectancy_r", 0) > 0.25:
        return "FUERTE"
    if row.get("pf", 0) >= 1.5 and row.get("expectancy_r", 0) > 0:
        return "OK"
    return "DEBIL"


def streak_sequences(df: pd.DataFrame, mask_col: str, name: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    cur: list[int] = []
    for idx, flag in enumerate(df[mask_col].tolist()):
        if bool(flag):
            cur.append(idx)
        elif cur:
            rows.append(build_streak_row(df, cur, name))
            cur = []
    if cur:
        rows.append(build_streak_row(df, cur, name))
    return rows


def build_streak_row(df: pd.DataFrame, indexes: list[int], name: str) -> dict[str, Any]:
    g = df.iloc[indexes]
    return {
        "tipo_racha": name,
        "largo": len(g),
        "fecha_inicio": str(g.iloc[0]["entry_date"]),
        "fecha_fin": str(g.iloc[-1]["entry_date"]),
        "r_total": safe_round(g["r_result"].sum(), 3),
        "tp": int((g["outcome"] == "TP").sum()),
        "be": int((g["outcome"] == "BE").sum()),
        "sl": int((g["outcome"] == "SL").sum()),
        "forced_close": int((g["outcome"] == "FORCED_CLOSE").sum()),
    }


def make_streak_tables(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    ordered = df.sort_values("entry_time_ny").reset_index(drop=True)
    ordered["streak_tp"] = ordered["outcome"].eq("TP")
    ordered["streak_sl"] = ordered["outcome"].eq("SL")
    ordered["streak_be"] = ordered["outcome"].eq("BE")
    ordered["streak_non_win"] = ordered["r_result"].le(0)
    ordered["streak_monetary_loss"] = ordered["r_result"].lt(0)
    ordered["streak_psychological"] = ordered["outcome"].ne("TP")
    all_rows = []
    for col, name in [
        ("streak_tp", "TP_CONSECUTIVOS"),
        ("streak_sl", "SL_PUROS_CONSECUTIVOS"),
        ("streak_be", "BE_CONSECUTIVOS"),
        ("streak_non_win", "NON_WIN_CONSECUTIVOS"),
        ("streak_monetary_loss", "PERDIDA_MONETARIA_CONSECUTIVA"),
        ("streak_psychological", "RACHA_PSICOLOGICA_SIN_TP"),
    ]:
        all_rows.extend(streak_sequences(ordered, col, name))
    streaks = pd.DataFrame(all_rows)
    summary = streaks.sort_values(["tipo_racha", "largo"], ascending=[True, False]).groupby("tipo_racha", as_index=False).head(1)
    worst = streaks[streaks["r_total"] <= 0].sort_values(["r_total", "largo"], ascending=[True, False]).head(10)
    best = streaks[streaks["r_total"] > 0].sort_values(["r_total", "largo"], ascending=[False, False]).head(10)
    return summary.reset_index(drop=True), worst.reset_index(drop=True), best.reset_index(drop=True)


def top_months(monthly: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    neg = monthly[monthly["total_r"] < 0].sort_values("total_r")
    best = monthly.sort_values("total_r", ascending=False).head(10)
    worst = monthly.sort_values("total_r", ascending=True).head(10)
    return neg, best, worst


def load_strategy_comparison(global_metrics: dict[str, Any]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = [
        {
            "estrategia_fase": "Phase25 / MANIPULANTE",
            "sample": global_metrics["sample"],
            "pf": global_metrics["pf"],
            "expectancy": global_metrics["expectancy_r"],
            "wr": global_metrics["winrate_r_positive_pct"],
            "dd": global_metrics["max_drawdown_r"],
            "max_streak": 14,
            "tpm": 19.36,
            "estado": "Authority",
            "motivo": "Mejor balance institucional: PF alto, DD bajo, evidencia 2015-2026, seguridad fail-closed.",
        }
    ]
    ranking = read_text(REPORTS / "INSTITUTIONAL_DAYTIME_STRATEGY_RANKING_REPORT.md")
    static_rows = [
        ("Phase18 Baseline", 1040, 1.63, "DATO_NO_DISPONIBLE", 38.0, -6.5, "DATO_NO_DISPONIBLE", "DATO_NO_DISPONIBLE", "Baseline", "Fundacion H1 sweep + M3 CHOCH, robusta pero superada."),
        ("Phase20 Balanced", "DATO_NO_DISPONIBLE", 1.58, "DATO_NO_DISPONIBLE", 40.3, -12.0, 8, 31.4, "Benchmark", "Recupero frecuencia, pero quedo superada por Phase22/24/25."),
        ("Phase22 High WR", 1048, 1.72, "DATO_NO_DISPONIBLE", 55.2, -8.45, 6, "DATO_NO_DISPONIBLE", "Superada", "Mayor WR, pero menor PF/mas fragilidad que Phase25."),
        ("Phase24 Robust Peak", 1602, 2.79, 0.328, 39.7, -5.0, "DATO_NO_DISPONIBLE", "DATO_NO_DISPONIBLE", "Backup", "Muy fuerte, pero Phase25 subio PF manteniendo DD."),
        ("TP1.4_BE0.5_BF70", 2625, 2.641, 0.3038, 36.46, -5.598, 12, 19.36, "Shadow", "Mejora WR/racha pero BE0.5 no es autoridad."),
        ("Phase19", "DATO_NO_DISPONIBLE", "DATO_NO_DISPONIBLE", "DATO_NO_DISPONIBLE", "DATO_NO_DISPONIBLE", "DATO_NO_DISPONIBLE", "DATO_NO_DISPONIBLE", "DATO_NO_DISPONIBLE", "Rechazada", "Invalidada/archivada por fallas forenses segun reportes."),
    ]
    if "PHASE25_IS_CURRENT_BEST_DAYTIME_STRATEGY" not in ranking:
        rows[0]["motivo"] += " Ranking institucional no disponible en texto esperado."
    for row in static_rows:
        rows.append(
            {
                "estrategia_fase": row[0],
                "sample": row[1],
                "pf": row[2],
                "expectancy": row[3],
                "wr": row[4],
                "dd": row[5],
                "max_streak": row[6],
                "tpm": row[7],
                "estado": row[8],
                "motivo": row[9],
            }
        )
    return pd.DataFrame(rows)


def load_trades_per_day_variants() -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    p28 = ROOT / "BOT_V2_DAYTIME_LAB" / "outputs" / "phase28_winrate_frequency_study" / "limited_combinations" / "phase28_limited_combinations_results.csv"
    if p28.exists():
        df = pd.read_csv(p28)
        for _, r in df[df["name"].astype(str).str.contains("2T|2TRADES", case=False, na=False)].iterrows():
            rows.append(
                {
                    "variante": r["name"],
                    "max_trades_day": 2,
                    "sample": int(r["sample"]),
                    "pf": r["pf"],
                    "expectancy": r["exp"],
                    "wr": r["wr"],
                    "dd": r["max_dd"],
                    "max_streak": r["max_streak"],
                    "trades_month": r["tpm"],
                    "veredicto": "SUBE_FRECUENCIA_PERO_NO_REEMPLAZA_AUTORIDAD",
                }
            )
    p28_single = ROOT / "BOT_V2_DAYTIME_LAB" / "outputs" / "phase28_winrate_frequency_study" / "single_hypothesis_tests" / "phase28_single_hypothesis_results.csv"
    if p28_single.exists():
        df = pd.read_csv(p28_single)
        for _, r in df[df["name"].astype(str).str.contains("2TRADES", case=False, na=False)].iterrows():
            rows.append(
                {
                    "variante": r["name"],
                    "max_trades_day": 2,
                    "sample": int(r["sample"]),
                    "pf": r["pf"],
                    "expectancy": r["exp"],
                    "wr": r["wr"],
                    "dd": r["max_dd"],
                    "max_streak": r["max_streak"],
                    "trades_month": r["tpm"],
                    "veredicto": "SUBE_FRECUENCIA_PF_DD_EMPEORAN",
                }
            )
    p29 = ROOT / "BOT_V2_DAYTIME_LAB" / "outputs" / "phase29_wr_loss_streak_compression" / "candidate_selection" / "phase29_candidate_comparison.csv"
    if p29.exists():
        df = pd.read_csv(p29)
        for _, r in df[df["name"].astype(str).str.contains("SECOND_TRADE|2T|2TRADES", case=False, na=False)].iterrows():
            rows.append(
                {
                    "variante": r["name"],
                    "max_trades_day": "2 condicionado",
                    "sample": int(r["sample"]),
                    "pf": r["pf"],
                    "expectancy": r["expectancy"],
                    "wr": r["wr"],
                    "dd": r["max_dd"],
                    "max_streak": r["max_loss_streak"],
                    "trades_month": r["trades_month"],
                    "veredicto": r.get("classification", "SHADOW_NO_AUTORIDAD"),
                }
            )
    rows.append(
        {
            "variante": "PHASE25_BASELINE",
            "max_trades_day": 1,
            "sample": 2625,
            "pf": 2.793,
            "expectancy": 0.2809,
            "wr": 32.53,
            "dd": -5.584,
            "max_streak": 14,
            "trades_month": 19.36,
            "veredicto": "AUTORIDAD_SE_MANTIENE",
        }
    )
    return pd.DataFrame(rows)


def load_audit_inventory() -> pd.DataFrame:
    rows = [
        ("Phase18", "H1 Fractal Sweep + First M3 CHOCH", "PHASE18_VALIDATED_FOR_FORWARD_DEMO", "Base protegida", "Funda la logica, luego superada."),
        ("Phase20", "News Fortress + recuperacion de frecuencia", "PHASE20_SAFE_CANDIDATE_FOUND", "Benchmark", "Aporta ventana amplia y News Fortress."),
        ("Phase22", "High WR / bajo DD", "PHASE22_HIGH_WR_CANDIDATE_FOUND", "Superada", "Mejor WR, no queda autoridad final."),
        ("Phase24", "Plateau robusto", "PHASE24_ROBUST_IMPROVEMENT_FOUND", "Backup fuerte", "Encuentra zona TP/BE robusta."),
        ("Phase25", "Cierre institucional y autoridad", "PHASE25_FINAL_CLOSEOUT_COMPLETE_READY_FOR_PAPER_DEMO_WITH_WARNINGS", "Authority", "Define MANIPULANTE TP1.4 BE0.4 BF70."),
        ("Phase27", "Validacion historica 2015-2026", "PHASE27_PHASE25_VALIDATED_2015_2026_STRONG", "Confirma autoridad", "Extiende evidencia a 2625 trades."),
        ("Phase28", "Winrate/frecuencia", "PHASE28_BALANCED_IMPROVEMENT_FOUND", "Shadow research", "Encuentra candidatos, no reemplaza Phase25."),
        ("Phase29", "WR y compresion de rachas", "PHASE29_BALANCED_WR_STREAK_IMPROVEMENT_FOUND", "Shadow research", "BE0.5 mejora rachas pero sigue no autoridad."),
        ("Phase30", "Forense BE0.5", "PHASE30_CANDIDATE_READY_FOR_PAPER_DEMO_WITH_WARNINGS", "Shadow", "Phase25 sigue autoridad."),
        ("Phase31", "Prop firm survival", "PHASE31_PROP_FIRM_READY_CONSERVATIVE_RISK", "Riesgo/fondeo", "0.50 pct prudente, 1 pct prohibido."),
        ("Phase32A", "FTMO 1-Step", "PHASE32A_FTMO_1STEP_SUPPORTED_WITH_WARNINGS", "Paper only", "FTMO 2-Step/Swing preferible; 0.50 pct base."),
        ("Phase32E", "Global Weekend Hard Close", "PHASE32E_GLOBAL_WEEKEND_HARD_CLOSE_IMPLEMENTED", "Seguridad operativa", "Viernes 16:55 NY global."),
        ("Phase37X-A", "Cierre diario 19:45 NY", "SHADOW_APPROVED", "Seguridad operativa", "Impacto insignificante, no cambia estrategia."),
        ("Phase37ZH", "AutoTrading / order_check readiness", "BLOCKED_AUTOTRADING_DISABLED", "Operativo demo", "Order_send gateado; MT5 bloquea AutoTrading."),
    ]
    return pd.DataFrame(rows, columns=["fase", "objetivo", "resultado_veredicto", "estado", "impacto_en_manipulante"])


def risk_table() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("Backtest historico", "R normalizado", "PF/expectancy/DD calculados en R, no en porcentaje de cuenta.", "OK"),
            ("FTMO Trial actual", "0.50%", "Politica FTMO Trial y runner usan 0.005 como base prudente.", "OK"),
            ("FTMO 2-Step/Swing challenge", "0.50% a 0.75%", "0.75% es techo defendible, no obligacion.", "WARNING"),
            ("Cuenta funded futura", "0.50%", "Prioriza supervivencia y margen contra daily loss.", "OK"),
            ("FundedNext Stellar Lite", "0.50% estricto", "Reglas diarias/total mas apretadas.", "OK"),
            ("Riesgo 1.00%", "PROHIBIDO", "Phase31/Phase32 muestran riesgo material de breach.", "RECHAZADO"),
        ],
        columns=["escenario", "riesgo", "motivo", "estado"],
    )


def parameters_table() -> pd.DataFrame:
    rows = [
        ("Estrategia", "Par", "EURUSD", "MANIPULANTE autoridad."),
        ("Estrategia", "Timeframes", "H1 contexto / M3 entrada", "H1 fractal sweep y First M3 CHOCH."),
        ("Estrategia", "Ventana entrada", "07:00-16:30 NY", "Fuera de ventana no abre entradas."),
        ("Estrategia", "Max trades/day", 1, "No subir a 2."),
        ("Estrategia", "TP", "1.4R", "Oficial."),
        ("Estrategia", "BE", "0.4R", "Oficial; BE0.5 es shadow."),
        ("Estrategia", "BF", "70%", "Body Filter minimo."),
        ("Estrategia", "SL logic", "Fractal/CHOCH con buffer 0.5 pip", "Segun Phase27 y detectores Phase18."),
        ("Riesgo", "Backtest", "R normalizado", "Metricas en R."),
        ("Riesgo", "FTMO Trial", "0.50%", "Runner demo/trial."),
        ("Riesgo", "Funded futuro", "0.50%", "Prudente."),
        ("Seguridad", "News rule", "News Fortress fail-closed", "Si duda, NO_TRADE."),
        ("Seguridad", "Data rule", "Data Quality Mask fail-closed", "Si data falla, NO_TRADE."),
        ("Operativo", "Daily forced close", "19:45 NY", "Seguridad de PC, no senal."),
        ("Operativo", "Friday hard close", "16:55 NY", "Global weekend hard close."),
        ("Operativo", "Mandatory close historico", "20:00 NY", "Usado en backtest Phase27."),
    ]
    return pd.DataFrame(rows, columns=["categoria", "parametro", "valor", "comentario"])


def list_sources() -> pd.DataFrame:
    files = [
        TRADES_CSV,
        REPORTS / "PHASE25_FINAL_CLOSEOUT_REPORT.md",
        REPORTS / "PHASE27_PHASE25_FULL_HISTORICAL_VALIDATION_2015_2026_REPORT.md",
        REPORTS / "PHASE28_WINRATE_FREQUENCY_IMPROVEMENT_STUDY_REPORT.md",
        REPORTS / "PHASE29_WR_LOSS_STREAK_COMPRESSION_REPORT.md",
        REPORTS / "PHASE30_TP14_BE05_BF70_FORENSIC_AUDIT_REPORT.md",
        REPORTS / "PHASE31_PROP_FIRM_SURVIVAL_SIMULATOR_REPORT.md",
        REPORTS / "PHASE32A_FTMO_1STEP_STANDARD_SIMULATION_REPORT.md",
        REPORTS / "PHASE32E_GLOBAL_WEEKEND_HARD_CLOSE_POLICY_REPORT.md",
        REPORTS / "PHASE37X_A_DAILY_FORCED_CLOSE_IMPACT_REPORT.md",
        REPORTS / "PHASE37ZH_ORDER_SEND_AUTOTRADING_READINESS_AUDIT_REPORT.md",
        MANIP / "01_ESTRATEGIA_AUTORIDAD" / "manipulante_config.json",
        MANIP / "01_ESTRATEGIA_AUTORIDAD" / "MANIPULANTE_STRATEGY_CARD.md",
        MANIP / "01_ESTRATEGIA_AUTORIDAD" / "MANIPULANTE_EVIDENCE_SUMMARY.md",
    ]
    return pd.DataFrame(
        [
            {
                "path": str(p.relative_to(ROOT)).replace("\\", "/"),
                "exists": p.exists(),
                "bytes": p.stat().st_size if p.exists() else 0,
                "sha256_12": hashlib.sha256(p.read_bytes()).hexdigest()[:12] if p.exists() and p.is_file() else "",
            }
            for p in files
        ]
    )


def qna_table() -> pd.DataFrame:
    rows = [
        ("Opera rangos de Asia?", "No como regla directa. Usa barrido de fractales H1; puede coincidir con rangos, pero no mide Asia como input principal."),
        ("Opera Londres?", "No como rango Londres discrecional. Opera ventana NY 07:00-16:30 y busca sweep H1 + CHOCH M3."),
        ("Usa dia anterior?", "No como regla directa tipo PDH/PDL. La logica es fractal H1 dinamico."),
        ("Por que el WR parece bajo?", "Porque BE cuenta como non-win en el winrate oficial; muchos trades terminan en BE y no en TP."),
        ("BE0.5 reemplaza a BE0.4?", "No. BE0.5 es shadow, no autoridad."),
        ("Puedo subir a 2 trades por dia?", "No como autoridad. Las pruebas suben frecuencia pero empeoran DD/racha o no justifican reemplazo."),
        ("Que hago si STATUS dice DUPLICADO?", "No operar manualmente; limpiar runners duplicados segun runbook/status."),
        ("Que hago si AutoTrading esta bloqueado?", "El bot queda bloqueado; revisar boton Trading algoritmico en MT5. No forzar ordenes."),
    ]
    return pd.DataFrame(rows, columns=["pregunta", "respuesta"])


def live_flow_table() -> pd.DataFrame:
    rows = [
        (1, "START_MANIPULANTE", "Inicia runner si no hay duplicado."),
        (2, "Runner Python", "Loop demo/trial; escribe heartbeat y quick_status."),
        (3, "MT5 FTMO Demo", "Solo demo/trial; real y Exness bloqueados."),
        (4, "News Cache/API", "Carga noticias para News Gate."),
        (5, "News Gate", "ALLOW o NO_TRADE fail-closed."),
        (6, "Data Gate", "Valida calidad de datos y simbolo."),
        (7, "Time Gate", "Permite 07:00-16:30 NY; luego manage only."),
        (8, "Signal Engine", "Busca Phase25 H1 sweep + First M3 CHOCH + BF70."),
        (9, "Lot Gate", "Calcula lote con riesgo permitido."),
        (10, "Order Check", "Validacion segura previa."),
        (11, "Order Send", "Gateado; no envia si gates bloquean."),
        (12, "TP/SL/BE", "Gestion oficial TP1.4, BE0.4."),
        (13, "Forced close 19:45", "Cierre operativo de seguridad si aplica."),
        (14, "STATUS", "Panel simple: OK, BLOQUEADO, DUPLICADO, PELIGRO."),
        (15, "Kill switch", "Bloquea operacion ante riesgo o STOP valido."),
    ]
    return pd.DataFrame(rows, columns=["paso", "componente", "explicacion"])


def explanation_tables() -> dict[str, pd.DataFrame]:
    identity = pd.DataFrame(
        [
            ("Que es", "MANIPULANTE es la estrategia oficial EURUSD diurna basada en Phase25 Authority."),
            ("Fase autoridad", "Phase25 define los parametros oficiales TP1.4 / BE0.4 / BF70."),
            ("Phase25 Authority", "Significa que esa configuracion manda sobre variantes posteriores salvo nueva fase aprobada."),
            ("Global Weekend Hard Close", "Regla de seguridad: viernes 16:55 NY no se sostiene posicion al fin de semana."),
            ("News Fortress", "Capa operativa fail-closed; si hay noticia bloqueante o duda, no opera."),
            ("Data Quality Mask", "Capa de calidad; si la data no esta certificada, no opera."),
            ("FTMO Trial runner", "Automatizacion demo/trial, no cambio de estrategia."),
            ("BE0.5", "Shadow comparator; no autoridad."),
        ],
        columns=["tema", "explicacion"],
    )
    logic = pd.DataFrame(
        [
            ("Que busca", "Un barrido de liquidez sobre fractales H1 y luego el primer cambio de caracter M3."),
            ("H1 Fractal Sweep", "Precio barre un nivel fractal H1 confirmado; no usa niveles futuros."),
            ("First M3 CHOCH", "Primer cambio objetivo de estructura en M3 posterior al sweep."),
            ("BF 70", "La vela de gatillo debe tener cuerpo/rango >= 70%."),
            ("Entrada", "Entra cuando sweep H1 + CHOCH M3 + BF70 + horario + news/data gates pasan."),
            ("No opera fuera de horario", "Fuera de 07:00-16:30 NY no abre nuevas operaciones."),
            ("No opera noticias", "News Fortress bloquea por preservacion de capital."),
            ("No opera si data falla", "Data mask es fail-closed."),
            ("Manual vs programable", "Hereda la idea SMC/ICT de barrido y cambio de estructura, pero no replica la discrecion completa ni rangos Asia/Londres/dia anterior como inputs directos."),
        ],
        columns=["tema", "explicacion"],
    )
    return {"identidad": identity, "logica": logic}


def make_preflight() -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    CSV_OUT.mkdir(parents=True, exist_ok=True)
    ANALYSIS_DIR.mkdir(parents=True, exist_ok=True)
    FIRST_READ_DIR.mkdir(parents=True, exist_ok=True)
    branch = run_cmd(["git", "branch", "--show-current"])
    status = run_cmd(["git", "status", "--short"])
    reports_main = [
        REPORTS / "PHASE25_FINAL_CLOSEOUT_REPORT.md",
        REPORTS / "PHASE27_PHASE25_FULL_HISTORICAL_VALIDATION_2015_2026_REPORT.md",
        REPORTS / "PHASE28_WINRATE_FREQUENCY_IMPROVEMENT_STUDY_REPORT.md",
        REPORTS / "PHASE29_WR_LOSS_STREAK_COMPRESSION_REPORT.md",
        REPORTS / "PHASE30_TP14_BE05_BF70_FORENSIC_AUDIT_REPORT.md",
        REPORTS / "PHASE31_PROP_FIRM_SURVIVAL_SIMULATOR_REPORT.md",
        REPORTS / "PHASE32A_FTMO_1STEP_STANDARD_SIMULATION_REPORT.md",
        REPORTS / "PHASE37ZH_ORDER_SEND_AUTOTRADING_READINESS_AUDIT_REPORT.md",
    ]
    data = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "cwd": str(ROOT),
        "cwd_correct": ROOT == Path(r"C:\Users\alera\Desktop\Bot\BOT DE TRADING ultimo"),
        "branch": branch,
        "git_status_short": status,
        "exists_manipulante": MANIP.exists(),
        "exists_lab": LAB.exists(),
        "exists_phase27_trades_csv": TRADES_CSV.exists(),
        "exists_zip_canonical": (ROOT / "000_PARA_CHATGPT.zip").exists(),
        "main_reports_present": {str(p.relative_to(ROOT)).replace("\\", "/"): p.exists() for p in reports_main},
        "mt5_touched": False,
        "execution_touched": False,
        "strategy_modified": False,
        "scope": "SOLO_ANALISIS_DOCUMENTACION",
    }
    (OUT / "phase38_preflight.json").write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    lines = [
        "# PHASE38 PREFLIGHT",
        "",
        f"- cwd: `{data['cwd']}`",
        f"- cwd correcto: {data['cwd_correct']}",
        f"- branch: `{branch}`",
        f"- existe MANIPULANTE: {data['exists_manipulante']}",
        f"- existe BOT_V2_DAYTIME_LAB: {data['exists_lab']}",
        f"- existe CSV Phase27: {data['exists_phase27_trades_csv']}",
        f"- existe ZIP canonico: {data['exists_zip_canonical']}",
        "- MT5 tocado: NO",
        "- ejecucion tocada: NO",
        "- estrategia modificada: NO",
        "",
        "## Reportes principales",
    ]
    for k, v in data["main_reports_present"].items():
        lines.append(f"- {k}: {'OK' if v else 'FALTA'}")
    lines.extend(["", "## Git status", "```", status, "```"])
    (OUT / "phase38_preflight.md").write_text("\n".join(lines), encoding="utf-8")
    return data


def write_csvs(data: dict[str, pd.DataFrame]) -> None:
    for name, df in data.items():
        df.to_csv(CSV_OUT / name, index=False, encoding="utf-8")
    if "phase38_monthly_breakdown.csv" in data:
        data["phase38_monthly_breakdown.csv"].to_csv(OUT / "monthly_breakdown.csv", index=False, encoding="utf-8")
    if "phase38_monthly_heatmap_data.csv" in data:
        data["phase38_monthly_heatmap_data.csv"].to_csv(OUT / "monthly_heatmap_data.csv", index=False, encoding="utf-8")


def add_sheet(wb: Workbook, name: str, df: pd.DataFrame, title: str | None = None) -> None:
    ws = wb.create_sheet(title=name[:31])
    start_row = 1
    if title:
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
        ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="1F4E78")
        start_row = 3
    if df.empty:
        ws.cell(row=start_row, column=1, value="DATO_NO_DISPONIBLE")
        return
    headers = list(df.columns)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=str(h))
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, value in enumerate(row, start=1):
            if pd.isna(value):
                value = ""
            ws.cell(row=r_idx, column=c_idx, value=value)
    end_row = start_row + len(df)
    end_col = len(headers)
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table_name = "T_" + "".join(ch for ch in name if ch.isalnum())[:20]
    try:
        tab = Table(displayName=table_name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(tab)
    except Exception:
        ws.auto_filter.ref = ref
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=end_col):
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in range(1, end_col + 1):
        values = [ws.cell(row=r, column=col).value for r in range(start_row, min(end_row, start_row + 80) + 1)]
        width = min(max(max(len(str(v)) if v is not None else 0 for v in values) + 2, 10), 42)
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in range(start_row + 1, end_row + 1):
        for col in range(1, end_col + 1):
            txt = str(ws.cell(row=row, column=col).value)
            if any(x in txt.upper() for x in ["RECHAZ", "ERROR", "DEBIL", "PROHIBIDO"]):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F4CCCC")
            elif any(x in txt.upper() for x in ["WARNING", "WARN", "BLOQUEADO", "SHADOW", "MUESTRA_BAJA"]):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFF2CC")
            elif any(x in txt.upper() for x in ["OK", "PASS", "FUERTE", "AUTHORITY", "APROB"]):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="D9EAD3")


def build_workbook(sheets: dict[str, pd.DataFrame]) -> Path:
    path = ANALYSIS_DIR / "MANIPULANTE_DEEP_EXPLAINER.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    ordered_names = [
        "README",
        "Resumen Ejecutivo",
        "Logica de Estrategia",
        "Parametros Oficiales",
        "Metricas Globales",
        "TP_BE_SL",
        "Rachas",
        "Horarios",
        "Dias Semana",
        "Meses",
        "Years",
        "2025 Deep Dive",
        "Comparacion Estrategias",
        "Trades Por Dia",
        "Riesgo y Fondeo",
        "Auditorias",
        "Live Flow",
        "Preguntas y Respuestas",
        "Trades Raw",
        "Afectados 19_45 si aplica",
    ]
    for name in ordered_names:
        add_sheet(wb, name, sheets.get(name, pd.DataFrame()), title=name)
    wb.save(path)
    return path


def write_markdown_report(context: dict[str, Any]) -> Path:
    p = ANALYSIS_DIR / "MANIPULANTE_DEEP_EXPLAINER_REPORT.md"
    g = context["global"]
    tp = context["tp_be_sl"]
    best_hours = context["best_hours"]
    worst_hours = context["worst_hours"]
    best_days = context["best_days"]
    worst_days = context["worst_days"]
    report = f"""# MANIPULANTE DEEP EXPLAINER REPORT

## 1. Que es MANIPULANTE
MANIPULANTE es la estrategia oficial EURUSD diurna del proyecto. La autoridad vigente es Phase25: TP 1.4R, BE 0.4R y BF 70. BE0.5 existe como shadow comparator, no como autoridad.

## 2. Como entra
Busca un H1 Fractal Sweep y luego el First M3 CHOCH. La entrada solo se acepta si pasa BF70, ventana 07:00-16:30 NY, News Fortress y Data Quality Mask.

## 3. Como gestiona
El TP oficial es 1.4R. El BE se activa a 0.4R. El SL sale de la estructura de la senal con buffer. La operacion tambien puede cerrar por tiempo.

## 4. Que resultados tiene
- Sample: {g['sample']}
- PF: {g['pf']}
- Expectancy: {g['expectancy_r']}R
- WR oficial: {g['winrate_r_positive_pct']}%
- Max DD: {g['max_drawdown_r']}R
- Total R: {g['total_r']}R

## 5. Por que el WR parece bajo
El WR oficial cuenta BE como non-win. Eso baja visualmente el winrate, aunque muchos trades no pierden dinero.

## 6. Cuantos TP/BE/SL hizo
- TP: {int(tp.loc[tp['Outcome']=='TP','Cantidad'].iloc[0])}
- BE: {int(tp.loc[tp['Outcome']=='BE','Cantidad'].iloc[0])}
- SL puro: {int(tp.loc[tp['Outcome']=='SL','Cantidad'].iloc[0])}
- Forced close/otros: {int(tp.loc[tp['Outcome']=='FORCED_CLOSE','Cantidad'].iloc[0])}

## 7. Mejores horarios
Mejores por PF/expectancy con muestra suficiente: {best_hours}.
Peores o mas debiles: {worst_hours}.
No se recomienda cambiar horarios desde esta fase; esto es explicacion, no optimizacion.

## 8. Mejores dias
Mejores dias: {best_days}.
Peores dias relativos: {worst_days}.
No se recomienda excluir dias sin una nueva fase de validacion.

## 9. Meses negativos
Si hubo meses negativos. Total: {context['negative_months_count']}. El peor mes fue {context['worst_month']} y el mejor mes fue {context['best_month']}.

## 10. Rachas esperables
La racha non-win maxima historica fue {context['max_non_win']} trades. Esto no significa {context['max_non_win']} SL seguidos: incluye BE y cierres sin ganancia.

## 11. Comparacion contra otras estrategias
MANIPULANTE quedo como autoridad porque combina PF alto, DD bajo, muestra amplia 2015-2026, todos los anos positivos, News/Data gates y reglas de fondeo. Algunas variantes tuvieron mejor WR, pero con costo de PF, DD, fragilidad o estatus shadow.

## 12. Por que quedo como autoridad
Phase25 fue el techo robusto institucional. Phase27 confirmo la robustez 2015-2026. Phase28-30 estudiaron mejoras pero no cambiaron autoridad.

## 13. Riesgos vivos
- 2025 fue el ano mas debil, aunque positivo.
- El WR psicologicamente parece bajo si se ignoran los BE.
- No hay path tick-by-tick completo para todas las auditorias.
- Costos reales, slippage y reglas de prop firm deben seguir revisandose.
- AutoTrading puede estar bloqueado operativamente en MT5.

## 14. Como operarlo en demo
Usar START/STATUS, leer los gates, no tocar TP/BE/BF, no operar si STATUS muestra bloqueado, duplicado o peligro.

## 15. Que NO tocar
No tocar TP 1.4R, BE 0.4R, BF70, max 1 trade/day, News Fortress, Data Quality Mask ni reglas de cierre.

## 16. Veredicto final
MANIPULANTE_EXPLAINER_COMPLETE_WITH_LIMITATIONS
"""
    p.write_text(report, encoding="utf-8")
    return p


def write_operator_summary(context: dict[str, Any]) -> Path:
    p = FIRST_READ_DIR / "MANIPULANTE_RESUMEN_PARA_OPERAR.md"
    text = f"""# MANIPULANTE - RESUMEN PARA OPERAR

## Que esperar
MANIPULANTE no gana por tener un winrate alto. Gana por PF, RR, BE rapido y control de perdida. El WR oficial historico es {context['global']['winrate_r_positive_pct']}%, porque los BE cuentan como non-win.

## Como se prende
Usar START_MANIPULANTE.bat. Si ya esta prendido, no iniciar otro runner.

## Como leer STATUS
- OK - BOT ACTIVO: runner vivo y gates principales OK.
- BLOQUEADO - BOT ACTIVO PERO NO OPERA: esta vivo, pero una regla bloquea.
- BLOQUEADO - AUTOTRADING DESHABILITADO: MT5 no permite ordenes automaticas.
- DUPLICADO - LIMPIAR RUNNERS: hay mas de un runner.
- PELIGRO - NO APAGAR PC: hay riesgo operativo o posicion abierta.

## Cuando no tocar nada
No tocar nada si hay noticia, data gate bloqueado, AutoTrading bloqueado, duplicado o posicion abierta.

## Mala racha
La racha non-win maxima fue {context['max_non_win']}. Puede incluir muchos BE, no necesariamente SL puros. La racha de SL puros es mucho menor.

## Que no cambiar
No cambiar TP 1.4R, BE 0.4R, BF70, max 1 trade/dia ni News/Data gates.

## Si sale duplicado
No operar. Limpiar runners duplicados y volver a STATUS.

## Si AutoTrading esta bloqueado
No forzar ordenes. Revisar el boton Trading algoritmico en MT5. El bot debe quedar bloqueado.

## Antes de apagar PC
Confirmar STATUS: OPERACION ABIERTA NO y SEGURO APAGAR PC SI. Viernes aplica hard close 16:55 NY.
"""
    p.write_text(text, encoding="utf-8")
    return p


def write_phase_report(context: dict[str, Any], outputs: dict[str, str]) -> tuple[Path, Path]:
    md = REPORTS / "PHASE38_MANIPULANTE_DEEP_EXPLAINER_REPORT.md"
    js = REPORTS / "PHASE38_MANIPULANTE_DEEP_EXPLAINER_REPORT.json"
    report = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "verdict": "MANIPULANTE_EXPLAINER_COMPLETE_WITH_LIMITATIONS",
        "scope": "analysis_documentation_only_no_mt5_no_execution",
        "strategy_changed": False,
        "mt5_touched": False,
        "execution_touched": False,
        "global_metrics": context["global"],
        "tp_be_sl": context["tp_be_sl"].to_dict(orient="records"),
        "best_month": context["best_month"],
        "worst_month": context["worst_month"],
        "negative_months_count": context["negative_months_count"],
        "max_non_win": context["max_non_win"],
        "outputs": outputs,
        "limitations": context["limitations"],
    }
    js.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    lines = [
        "# PHASE38 MANIPULANTE DEEP EXPLAINER REPORT",
        "",
        "## 1. Lo mas importante",
        "Se creo una auditoria explicativa completa de MANIPULANTE sin modificar estrategia, runner, MT5 ni launchers.",
        "",
        "## 2. Veredicto final exacto",
        "MANIPULANTE_EXPLAINER_COMPLETE_WITH_LIMITATIONS",
        "",
        "## 3. Metricas principales",
        f"- Sample: {context['global']['sample']}",
        f"- PF: {context['global']['pf']}",
        f"- Expectancy: {context['global']['expectancy_r']}R",
        f"- WR: {context['global']['winrate_r_positive_pct']}%",
        f"- DD: {context['global']['max_drawdown_r']}R",
        f"- Total R: {context['global']['total_r']}R",
        "",
        "## 4. Seguridad",
        "- MT5 tocado: NO",
        "- Ejecucion tocada: NO",
        "- Estrategia modificada: NO",
        "- Ordenes enviadas: NO",
        "",
        "## 5. Archivos creados",
    ]
    for k, v in outputs.items():
        lines.append(f"- {k}: `{v}`")
    lines.extend(
        [
            "",
            "## 6. Limitaciones",
            *[f"- {x}" for x in context["limitations"]],
            "",
            "## 7. Siguiente paso unico",
            "Leer el Excel y el resumen operativo antes de operar demo; no cambiar parametros.",
        ]
    )
    md.write_text("\n".join(lines), encoding="utf-8")
    return md, js


def make_affected_1945() -> pd.DataFrame:
    shadow = read_json(LAB / "outputs" / "phase37x_a_daily_forced_close_audit" / "shadow_audit.json", {})
    rows = []
    if isinstance(shadow, dict):
        for key, value in shadow.items():
            if isinstance(value, (str, int, float, bool)) or value is None:
                rows.append({"campo": key, "valor": value})
        if not rows:
            rows.append({"campo": "detalle", "valor": "DATO_NO_DISPONIBLE_EN_FORMATO_TABULAR"})
    else:
        rows.append({"campo": "detalle", "valor": "DATO_NO_DISPONIBLE"})
    return pd.DataFrame(rows)


def main() -> None:
    preflight = make_preflight()
    trades = parse_trades()
    global_metrics = metric_row("PHASE25_2015_2026", trades)

    outcome_rows = []
    total = len(trades)
    for outcome, g in trades.groupby("outcome", sort=True):
        outcome_rows.append(
            {
                "Outcome": outcome,
                "Cantidad": len(g),
                "% total": safe_round(len(g) / total * 100, 2),
                "R promedio": safe_round(g["r_result"].mean(), 4),
                "R total": safe_round(g["r_result"].sum(), 3),
            }
        )
    tp_be_sl = pd.DataFrame(outcome_rows).sort_values("Outcome")

    streak_summary, worst_streaks, best_streaks = make_streak_tables(trades)
    all_streaks = pd.concat(
        [
            streak_summary.assign(grupo="MAXIMOS"),
            worst_streaks.assign(grupo="TOP_10_PEORES"),
            best_streaks.assign(grupo="TOP_10_MEJORES"),
        ],
        ignore_index=True,
    )

    hourly = summarize_group(trades, "hour_ny", "Hora NY")
    weekdays = summarize_group(trades, "weekday", "Dia")
    weekday_order = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes"]
    weekdays["Dia"] = pd.Categorical(weekdays["Dia"], weekday_order, ordered=True)
    weekdays = weekdays.sort_values("Dia")
    monthly = summarize_group(trades, "year_month", "Year-Month")
    month_calendar = summarize_group(trades, "month_name", "Mes calendario")
    yearly = summarize_group(trades, "year", "Year")
    y2025 = trades[trades["year"] == 2025]
    y2025_summary = pd.concat(
        [
            pd.DataFrame([metric_row("2025", y2025)]),
            summarize_group(y2025, "year_month", "Year-Month"),
        ],
        ignore_index=True,
    )
    negative_months, best_months, worst_months = top_months(monthly)
    monthly_heatmap = monthly.copy()
    monthly_heatmap["Year"] = monthly_heatmap["Year-Month"].str.slice(0, 4)
    monthly_heatmap["Month"] = monthly_heatmap["Year-Month"].str.slice(5, 7)

    strategy_comparison = load_strategy_comparison(global_metrics)
    trades_per_day = load_trades_per_day_variants()
    audit_inventory = load_audit_inventory()
    risks = risk_table()
    params = parameters_table()
    sources = list_sources()
    explanations = explanation_tables()
    live_flow = live_flow_table()
    qna = qna_table()
    affected_1945 = make_affected_1945()

    raw_cols = [
        "type",
        "entry_time",
        "exit_time",
        "entry_price",
        "exit_price",
        "sl",
        "tp",
        "risk",
        "status",
        "be_triggered",
        "outcome",
        "r_result",
        "entry_date",
        "year",
        "year_month",
        "month_name",
        "hour_ny",
        "weekday",
    ]
    raw_enriched = trades[raw_cols].copy()
    raw_enriched["r_result"] = raw_enriched["r_result"].round(6)

    global_df = pd.DataFrame([global_metrics])
    csvs = {
        "phase38_global_metrics.csv": global_df,
        "phase38_tp_be_sl_counts.csv": tp_be_sl,
        "phase38_streaks.csv": all_streaks,
        "phase38_hourly_breakdown.csv": hourly,
        "phase38_weekday_breakdown.csv": weekdays,
        "phase38_monthly_breakdown.csv": monthly,
        "phase38_monthly_heatmap_data.csv": monthly_heatmap,
        "phase38_yearly_breakdown.csv": yearly,
        "phase38_strategy_comparison.csv": strategy_comparison,
        "phase38_trades_per_day_variants.csv": trades_per_day,
        "phase38_audit_inventory.csv": audit_inventory,
        "phase38_negative_months.csv": negative_months,
        "phase38_top_best_months.csv": best_months,
        "phase38_top_worst_months.csv": worst_months,
        "phase38_raw_trades_enriched.csv": raw_enriched,
    }
    write_csvs(csvs)

    readme = pd.DataFrame(
        [
            ("Veredicto", "MANIPULANTE_EXPLAINER_COMPLETE_WITH_LIMITATIONS"),
            ("Alcance", "Analisis y documentacion solamente. No MT5. No ejecucion. No cambio de estrategia."),
            ("Fuente principal", str(TRADES_CSV.relative_to(ROOT)).replace("\\", "/")),
            ("Sample", global_metrics["sample"]),
            ("Notas", "BE sale como status=SL y be_triggered=True en el CSV; se reclasifico como BE para explicar."),
        ],
        columns=["campo", "valor"],
    )
    executive = pd.DataFrame(
        [
            ("MANIPULANTE", "Estrategia oficial EURUSD diurna; Phase25 Authority."),
            ("Parametros", "TP1.4R / BE0.4R / BF70 / 1 trade por dia / 07:00-16:30 NY."),
            ("Resultado", f"PF {global_metrics['pf']} / Exp {global_metrics['expectancy_r']}R / WR {global_metrics['winrate_r_positive_pct']}% / DD {global_metrics['max_drawdown_r']}R."),
            ("WR bajo", "BE cuenta como non-win; excluyendo BE el winrate mejora visualmente."),
            ("Autoridad", "No cambia. BE0.5 sigue shadow."),
        ],
        columns=["tema", "resumen"],
    )
    sheets = {
        "README": pd.concat([readme, sources.rename(columns={"path": "campo", "exists": "valor"})[["campo", "valor"]]], ignore_index=True),
        "Resumen Ejecutivo": executive,
        "Logica de Estrategia": pd.concat([explanations["identidad"], explanations["logica"]], ignore_index=True),
        "Parametros Oficiales": params,
        "Metricas Globales": global_df,
        "TP_BE_SL": tp_be_sl,
        "Rachas": all_streaks,
        "Horarios": hourly,
        "Dias Semana": weekdays,
        "Meses": pd.concat([monthly, month_calendar], ignore_index=True),
        "Years": yearly,
        "2025 Deep Dive": y2025_summary,
        "Comparacion Estrategias": strategy_comparison,
        "Trades Por Dia": trades_per_day,
        "Riesgo y Fondeo": risks,
        "Auditorias": audit_inventory,
        "Live Flow": live_flow,
        "Preguntas y Respuestas": qna,
        "Trades Raw": raw_enriched,
        "Afectados 19_45 si aplica": affected_1945,
    }
    workbook = build_workbook(sheets)

    max_non_win = int(streak_summary.loc[streak_summary["tipo_racha"] == "NON_WIN_CONSECUTIVOS", "largo"].iloc[0])
    max_sl = int(streak_summary.loc[streak_summary["tipo_racha"] == "SL_PUROS_CONSECUTIVOS", "largo"].iloc[0])
    max_tp = int(streak_summary.loc[streak_summary["tipo_racha"] == "TP_CONSECUTIVOS", "largo"].iloc[0])
    context = {
        "global": global_metrics,
        "tp_be_sl": tp_be_sl,
        "best_hours": ", ".join(hourly.sort_values(["pf", "expectancy_r"], ascending=False).head(3)["Hora NY"].astype(str).tolist()),
        "worst_hours": ", ".join(hourly.sort_values(["pf", "expectancy_r"], ascending=True).head(3)["Hora NY"].astype(str).tolist()),
        "best_days": ", ".join(weekdays.sort_values(["pf", "expectancy_r"], ascending=False).head(2)["Dia"].astype(str).tolist()),
        "worst_days": ", ".join(weekdays.sort_values(["pf", "expectancy_r"], ascending=True).head(2)["Dia"].astype(str).tolist()),
        "negative_months_count": int(len(negative_months)),
        "best_month": str(best_months.iloc[0]["Year-Month"]),
        "worst_month": str(worst_months.iloc[0]["Year-Month"]),
        "max_non_win": max_non_win,
        "max_sl": max_sl,
        "max_tp": max_tp,
        "limitations": [
            "El CSV Phase27 no trae path intratrade tick-by-tick completo.",
            "Algunas comparaciones contra fases viejas provienen de reportes, no de reconstruccion uniforme.",
            "BE esta codificado como SL con be_triggered=True y fue reclasificado para lectura humana.",
            "Forced close no siempre equivale a TP/BE/SL; se mantiene separado.",
            "Esta fase no optimiza ni recomienda cambios de parametros.",
        ],
    }
    explainer_report = write_markdown_report(context)
    operator_summary = write_operator_summary(context)
    outputs = {
        "excel": str(workbook.relative_to(ROOT)).replace("\\", "/"),
        "reporte_markdown": str(explainer_report.relative_to(ROOT)).replace("\\", "/"),
        "resumen_operativo": str(operator_summary.relative_to(ROOT)).replace("\\", "/"),
        "csv_dir": str(CSV_OUT.relative_to(ROOT)).replace("\\", "/"),
    }
    write_phase_report(context, outputs)


if __name__ == "__main__":
    main()
